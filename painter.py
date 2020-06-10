import argparse
import os
import sys

import cv2
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from progress.bar import PixelBar


def rgb2hex(r, g, b):
    return "{:02x}{:02x}{:02x}".format(r, g, b)


def paint_spreadsheet(y_max, x_max, image, filename):

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Crafted With Love"

    with PixelBar("Painting Canvas", max=x_max) as bar:
        for col in range(1, x_max):
            for row in range(1, y_max):

                row_r = row * 3
                row_g = row_r + 1
                row_b = row_r + 2

                cell_r = ws1.cell(row_r, col)
                cell_g = ws1.cell(row_g, col)
                cell_b = ws1.cell(row_b, col)

                colors = image[row - 1][col - 1]

                cell_r.fill = PatternFill(start_color=rgb2hex(colors[2], 0, 0), fill_type="solid")
                cell_g.fill = PatternFill(start_color=rgb2hex(0, colors[1], 0), fill_type="solid")
                cell_b.fill = PatternFill(start_color=rgb2hex(0, 0, colors[0]), fill_type="solid")

                # ws1.row_dimensions[row_r].height = 3.75
                # ws1.row_dimensions[row_g].height = 3.75
                # ws1.row_dimensions[row_b].height = 3.75
                # ws1.column_dimensions[get_column_letter(col)].width = 2

            bar.next()
    wb.save(filename=f"{filename}.xlsx")


if __name__ == "__main__":

    parser = argparse.ArgumentParser()
    parser.add_argument("--image", help="Path to image file", required=True)
    parser.add_argument(
        "--orignal",
        action="store_true",
        help="Keep the orignal size of the image.Without this image would be scaled down",
    )
    parser.add_argument(
        "--tiny",
        action="store_true",
        help="Scale down image to 12 percent.Usefull while dealing with very high resolution images",
    )

    image_path = parser.parse_args().image
    orignal = parser.parse_args().orignal
    tiny = parser.parse_args().tiny

    image = cv2.imread(image_path)

    if not orignal and not tiny:
        image = cv2.resize(image, (0, 0,), fx=0.35, fy=0.35)
    elif tiny and not orignal:
        image = cv2.resize(image, (0, 0,), fx=0.12, fy=0.12)

    paint_spreadsheet(len(image), len(image[0]), image, os.path.basename(image_path).split(".")[0])
